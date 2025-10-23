import requests
import json
import csv
import os
import platform
import subprocess
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

CONFIG_FILE = "config.json"

# ------------------ JSON FLATTENING ------------------
def flatten_json(data):
    flat = {}
    def recurse(t, parent_key=''):
        if isinstance(t, dict):
            for k, v in t.items():
                recurse(v, f"{parent_key}.{k}" if parent_key else k)
        elif isinstance(t, list):
            for i, v in enumerate(t):
                recurse(v, f"{parent_key}[{i}]")
        else:
            flat[parent_key] = t
    recurse(data)
    return flat

# ------------------ CONFIG ------------------
def save_config(config):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    return None

# ------------------ EXPORTS ------------------
def export_to_text(data_list, filename):
    with open(filename, "w", encoding="utf-8") as f:
        for item in data_list:
            json.dump(item, f, ensure_ascii=False, indent=2)
            f.write("\n\n")
    print(f"Data exported to {filename}")

def export_to_csv(data_list, filename):
    flat_data = [flatten_json(d) for d in data_list]
    keys = sorted({k for d in flat_data for k in d.keys()})
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(flat_data)
    print(f"Data exported to {filename}")

def export_to_excel(data_list, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    
    flat_data = [flatten_json(d) for d in data_list]
    keys = sorted({k for d in flat_data for k in d.keys()})
    ws.append(keys)
    
    for item in flat_data:
        ws.append([item.get(k, "") for k in keys])
    
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    wb.save(filename)
    print(f"Data exported to {filename}")

# ------------------ OPEN FILE ------------------
def open_file(filename):
    system_platform = platform.system()
    try:
        if system_platform == "Darwin":
            subprocess.call(["open", filename])
        elif system_platform == "Windows":
            os.startfile(filename)
        elif system_platform == "Linux":
            subprocess.call(["xdg-open", filename])
        else:
            print(f"Automatic open not supported on this OS: {system_platform}")
    except Exception as e:
        print(f"Could not open file automatically: {e}")

# ------------------ AUTO-DETECT FORMAT ------------------
def detect_best_format(data_list):
    if not data_list:
        return "txt"
    try:
        flat = flatten_json(data_list[0])
        keys = flat.keys()
        if len(keys) > 3 and all(isinstance(d, dict) for d in data_list):
            return "csv"
    except Exception:
        pass
    return "txt"

# ------------------ MAIN EXTRACTION ------------------
def run_extraction(config):
    urls = config.get("urls", [])
    export_format = config.get("export_format", "auto")

    data_list = []
    for url in urls:
        print(f"Processing: {url}")
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            if isinstance(data, list):
                data_list.extend(data)
            else:
                data_list.append(data)
        except Exception as e:
            print(f"Error fetching {url}: {e}")

    # Auto-detect format if needed
    if export_format == "auto":
        detected = detect_best_format(data_list)
        print(f"\nAuto-detected best format: {detected.upper()}")
        confirm = input(f"Proceed with {detected.upper()}? (Y/n): ").strip().lower()
        if confirm == "n":
            print("\nSelect preferred export format:")
            print("1. Text (.txt)")
            print("2. CSV (.csv)")
            print("3. Excel (.xlsx)")
            choice = input("Enter choice (1/2/3): ").strip()
            formats = {"1": "txt", "2": "csv", "3": "xlsx"}
            export_format = formats.get(choice, detected)
        else:
            export_format = detected

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"data_output_{timestamp}.{export_format}"

    if export_format == "txt":
        export_to_text(data_list, output_file)
    elif export_format == "csv":
        export_to_csv(data_list, output_file)
    elif export_format == "xlsx":
        export_to_excel(data_list, output_file)

    print(f"\nSUMMARY: {len(urls)} URLs processed | {len(data_list)} successful | Process completed!")

    open_now = input("Would you like to open the exported file now? (Y/n): ").strip().lower()
    if open_now == "y":
        open_file(output_file)

# ------------------ UI ------------------
def main():
    print("=" * 60)
    print("        AUTOMATED DATA EXTRACTOR - CONFIGURATION TOOL")
    print("=" * 60)

    config = None
    if os.path.exists(CONFIG_FILE):
        print(f"\nDetected existing configuration file: {CONFIG_FILE}")
        use_existing = input("Would you like to use it? (Y/n): ").strip().lower()
        if use_existing == "y":
            config = load_config()

    if not config:
        urls = input("\nEnter URLs (comma separated): ").strip().split(",")
        print("\nSelect export format:")
        print("1. Text (.txt)")
        print("2. CSV (.csv)")
        print("3. Excel (.xlsx)")
        print("4. Auto-detect best format")
        choice = input("Enter choice (1/2/3/4): ").strip()

        formats = {"1": "txt", "2": "csv", "3": "xlsx", "4": "auto"}
        export_format = formats.get(choice, "auto")
        config = {"urls": [u.strip() for u in urls if u.strip()], "export_format": export_format}

        save_pref = input("Save this configuration for next time? (Y/n): ").strip().lower()
        if save_pref == "y":
            save_config(config)
            print(f"Configuration saved to {CONFIG_FILE}")

    print("\nStarting extraction process...")
    run_extraction(config)

# ------------------ ENTRY ------------------
if __name__ == "__main__":
    main()
